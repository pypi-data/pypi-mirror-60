# Copyright (c) 2019-2020 Kevin Crouse
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#
# @license: http://www.apache.org/licenses/LICENSE-2.0
# @author: Kevin Crouse (krcrouse@gmail.com)

from .base import SourceBase
import os

class XLSX(SourceBase):

    def __init__(self, *args, **kwargs):
        self._iterator = None
        self.workbook = None
        super().__init__(*args, **kwargs)

    def open_stream(self):
        import openpyxl
        self.workbook = openpyxl.load_workbook(os.path.expanduser(self.filename), read_only=True, data_only=True)
        self._worksheets = self.workbook.sheetnames
        if self.worksheet:
            self.xlsheet = self.workbook[self.worksheet]
        else:
            self.xlsheet = self.workbook.worksheets[0] ### this could be self.workbook.active, but that has given weird results
            self.worksheet = self.xlsheet.title
        self._iterator = iter(self.xlsheet.rows)

    @property
    def iterator(self):
        if not self._iterator:
            self.open()
        return(self._iterator)

    @property
    def worksheets(self):
        if not self.workbook:
            self.open()
        return(self._worksheets)


    @property
    def worksheet(self): return(self._worksheet)

    @worksheet.setter
    def worksheet(self, worksheet):
        self._worksheet = worksheet
        self.xlsheet = self.workbook[worksheet]

    def skip_rows(self, count):
        skipped = []
        for i in range(count):
            row = next(self.iterator)
            skipped.append(list(map(lambda cell: cell.value, row)))
        return(skipped)

    def next_row(self):
        hasText = False
        while True:
            try:
                rowcontainer = next(self.iterator)
            except StopIteration:
                self.workbook.close()
                raise(StopIteration)

            row = []
            cellcount = 0
            for cell in rowcontainer:
                cellcount += 1
                v = cell.value
                if type(v) is str:
                    if self.nonemptyre.search(v):
                        hasText = cellcount
                    else:
                        v = None
                elif v is not None:
                    hasText = cellcount
                else:
                    v = None
                # add the cell value to the row
                row.append(v)
            if hasText is not False:
                return(row[:hasText])

    @classmethod
    def export_to(self, matrix, filename, topmatter=None, autosize=None):
        """The export_to class method to export a matrixb matrix to a modern Excel ('.xlsx') file.

        Args:
            matrix (matrixb.Matrix): The Matrix object to export.
            filename (str): The full path to send the file.
            topmatter (list|str, optional): Lines to appear above the exported table.
            autosize (bool): TODO.
        """
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        irow = 0
        # handle topmatter
        if topmatter:
            if type(topmatter) in (list, tuple):
                for rowdata in topmatter:
                    ws.append(rowdata)
                    irow += 1
            else:
                ws.append([topmatter])
                irow += 1
        # handle column headers, with some formatting

        headerborder = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thick',color='000000'))
        headerfont = openpyxl.styles.Font(bold=True, size=12)
        for icol in range(len(matrix.columns)):
            cell = ws.cell(irow+1, icol+1)
            cell.value = matrix.columns[icol]
            cell.font = headerfont
            cell.border = headerborder
        irow += 1

        for rowdata in matrix:
            ws.append(rowdata)
        wb.save(filename)

    def __getstate__(self):
        """ To pickle/serialze the csv source, we delete the stream and filehandle - this will allow future restored objects to get things like the source filename, but attempts to access the source object will fail. """
        state = super().__getstate__()
        del state['workbook']
        del state['xlsheet']
        del state['_iterator']
        return(state)
