# encoding: utf-8

import json
import logging_helper
from openpyxl import load_workbook

logging = logging_helper.setup_logging()


def cell_value(cell,
               strip):
    if strip:
        try:
            return cell.value.strip()
        except AttributeError:
            pass
    return cell.value


def get_col_dicts(worksheet,
                  strip):

    rows = [row for row in worksheet.rows]
    if not rows:
        return []

    keys = [cell_value(cell=cell,
                       strip=strip)
            for cell in rows.pop(0)]

    return [{k: v for k, v in zip(keys,
                                  [cell_value(cell=cell,
                                              strip=strip)
                                   for cell in row])}
            for row in rows]


def get_col_lists(worksheet,
                  header_row=0):

    rows = [row for row in worksheet.rows]
    if not rows:
        return []
    row_count = len(rows)
    col_count = max([len(row) for row in rows])

    columns = [[None for _ in range(row_count)] for _ in range(col_count)]
    for row_num, row in enumerate(rows):
        for col_num, cell in enumerate(row):
            columns[col_num][row_num] = cell.value

    col_dicts = {col[header_row]: col[header_row+1:]
                 for col in columns
                 if col[header_row] is not None}

    return col_dicts


def workbook_to_json(filename,
                     strip=True):
    workbook = load_workbook(filename=filename,
                             read_only=True)

    tables = {worksheet_name: get_col_dicts(worksheet=workbook[worksheet_name],
                                       strip=strip)
              for worksheet_name in workbook.sheetnames}

    # Strip out empty sheets
    tables = {table: items
              for table, items in iter(tables.items())
              if items}

    workbook.close()

    if len(tables) > 1:
        return json.dumps(tables)
    # Only 1 sheet/table.
    return json.dumps(list(tables.values())[0])

