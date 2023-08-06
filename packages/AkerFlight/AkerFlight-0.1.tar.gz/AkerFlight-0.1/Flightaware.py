from selenium import webdriver #Selenium powers the autobrowser
from selenium.webdriver.common.keys import Keys #Allows sending of keyboard presses to HTML elements
import time #Allows use of the sleep function
import openpyxl #Allows Python to read Excel files
from openpyxl import load_workbook #Allows Python to load Excel workbooks.
from datetime import date #Obtains system date

#A set is used instead of a list to remove duplicates, and order is uneccesarry.
finallist = {""}
#

#Takes the name of the airline, and shorts it to the IATA code.
def condenser(airline,flightnum):
    if flightnum != 'none':
        #This is not a complete list of airlines, will need ammending
        airline = airline.replace('American Airlines', 'AA')
        airline = airline.replace('Alaska Airlines', 'AS')
        airline = airline.replace('United Airlines', 'UA')
        airline = airline.replace('Delta Air Lines', 'DL')
        airline = airline.replace('Jetblue Airways Corporation', 'B6')
        airline = airline.replace('Southwest Airlines', 'WN')
        airline = airline.replace('Aer Lingus', 'EL')
        airline = airline.replace('Frontier Airlines Inc.', 'F9')
        airline = airline.replace('Virgin Atlantic Airways', 'VS')
        airline = airline.replace('Hawaiian Airlines', 'HA')
        airline = airline.replace('British Airways', 'BA')
        
        #
        
        #Condenses the airline and the flight number by adding them together into a string.
        condensed = airline+str(flightnum)
        #
        
        #TravelTracker exports airlines and trainlines as the same category. This wil remove amtrak entries from the set
        if 'Amtrak' not in condensed:
            finallist.add(condensed)
        #
#

#Loads the Excel doc, sets up variables.
workbook = load_workbook(filename="Flights.xlsx")
workbook.sheetnames
sheet = workbook.active
num = 1
Inbounddate = ""
Outbounddate= ""
today = date.today()
day = today.day
maximum = sheet.max_row
#

for num in range(1,maximum):
    #Column 2 is the Inbound date, Column 4 is the inbound Airline, Column 1 is Inbound Flight Number.
    Inbounddate = sheet.cell(row=num, column=2).value
    Inboundtime = str(sheet.cell(row=num, column = 3).value)
    #
    
    #Column 6 is the Outbound Date, and Column 8 is the Outbound Airline, Column 5 is the Outbound Flight number
    Outbounddate = sheet.cell(row=num, column =6).value
    Outboundtime = str(sheet.cell(row=num, column =7).value)
    #
    
    #Checks to see if the "day" in today's date is in the Flight Date.
    if str(day) in str(Inbounddate) and int(Inboundtime[0:2]) <=16:
        condenser((sheet.cell(row=num, column = 4).value),(sheet.cell(row=num, column=1).value))
    if str(day) in str(Outbounddate) and int(Outboundtime[0:2]) <=16:
        condenser((sheet.cell(row=num, column = 8).value),(sheet.cell(row=num, column=5).value))
    #
print(finallist)

# Start the autobrowser and then go to the login page of Flightaware
driver = webdriver.Chrome()
driver.get("https://flightaware.com/account/session")

#Targets the username and password boxes on the webpage and inputs the credentials for the account, then submits them.
username = driver.find_element_by_name('flightaware_username')
username.send_keys("")
password = driver.find_element_by_name('flightaware_password')
password.send_keys("")
password.send_keys(Keys.ENTER)
#

#Goes to the flight tracking management page, then targets the box to add flights
driver.get("https://flightaware.com/me/manage")
aircraft = driver.find_element_by_id('add_ident')
#

#Enters flight numbers into box, presses enter to submit them.
num = 0
for x in finallist:
    aircraft.send_keys(x)
    aircraft.send_keys(Keys.ENTER)
    time.sleep(1)
    num+=1
print('Completed '+ num + ' flights' )

