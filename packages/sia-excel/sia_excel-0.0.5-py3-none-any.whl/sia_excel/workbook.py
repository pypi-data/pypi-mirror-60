import enum
from pathlib import Path
from copy import copy
from typing import Iterable, Union, List

import pandas as pd
from pandas.core.indexes.frozen import FrozenList
from openpyxl import Workbook as _Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils.cell import (
    get_column_letter,
    column_index_from_string,
    coordinate_from_string,
)
from openpyxl.utils.dataframe import dataframe_to_rows


class CellFont(enum.auto):
    TABLE_TITLE = Font(name='Calibri', size='12', italic=True)
    TABLE_HEADER = Font(name='Calibri', size=10, bold=True)
    TABLE_DATA = copy(TABLE_HEADER)
    TABLE_DATA.bold = False


class Workbook:

    '''SIA layered customizations of openpyxl Workbook. '''

    def __init__(self, fp: str):
        self.pth = Path(fp).resolve()
        if self.pth.resolve().exists():
            self._workbook = load_workbook(fp)
        else:
            self._workbook = _Workbook()

    def close(self):
        self._workbook.save(self.pth)

    def write(
        self,
        data: pd.DataFrame,
        ws_name: str,
        node: str,
        header: bool = True,
        index: bool = False,
        cell_format: Union[None, List[CellFont]] = None,
    ) -> None:
        wb = self._workbook
        if ws_name in wb.sheetnames:
            ws = wb[ws_name]
        else:
            ws = wb.create_sheet(ws_name)
        height, width = data.shape
        if index:
            # Hack for adding index without adding a blank row after header
            width += 1
            data.insert(0, '', data.index)
        node_col, node_row = coordinate_from_string(node)
        end_col = get_column_letter(column_index_from_string(node_col) + width)
        end_row = node_row + height
        xl_range = ws[f'{node}:{end_col}{end_row}']
        df_rows = dataframe_to_rows(data, index=False, header=header)
        for df_row, xl_row in zip(df_rows, xl_range):
            for df_data, xl_cell in zip(df_row, xl_row):
                xl_cell.value = df_data
